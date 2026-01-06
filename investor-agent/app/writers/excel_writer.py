"""
Excel Writer.

Writes investment analysis results to Excel files.
"""

from typing import Dict, Any
import os
from datetime import datetime


class ExcelWriter:
    """
    Writes investment analysis results to Excel files.
    
    Creates formatted Excel workbooks with analysis data and charts.
    """
    
    def __init__(self, config):
        """
        Initialize the Excel writer.
        
        Args:
            config: Configuration object with output settings
        """
        self.config = config
        self.output_dir = config.output_dir
        
        # Ensure output directory exists
        os.makedirs(self.output_dir, exist_ok=True)
    
    def write(self, results: Dict[str, Any]) -> str:
        """
        Write investment analysis results to an Excel file.
        
        Args:
            results: Analysis results dictionary
        
        Returns:
            Path to the created Excel file
        """
        ticker = results.get("ticker", "UNKNOWN")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{ticker}_analysis_{timestamp}.xlsx"
        filepath = os.path.join(self.output_dir, filename)
        
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            self._write_summary_sheet(ws_summary, results)
            
            # Metrics sheet
            ws_metrics = wb.create_sheet("Metrics")
            self._write_metrics_sheet(ws_metrics, results)
            
            # Scores sheet
            ws_scores = wb.create_sheet("Scores")
            self._write_scores_sheet(ws_scores, results)
            
            # Save workbook
            wb.save(filepath)
            
            return filepath
            
        except ImportError:
            # Fallback to simple CSV-like format if openpyxl not available
            print("Warning: openpyxl not installed. Creating simple text file. Install with: pip install openpyxl")
            return self._write_simple_format(filepath.replace('.xlsx', '.txt'), results)
        except Exception as e:
            raise RuntimeError(f"Failed to write Excel file: {e}")
    
    def _write_summary_sheet(self, ws, results: Dict[str, Any]):
        """Write summary information to worksheet."""
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Title
        ws['A1'] = "Investment Analysis Summary"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Company info
        ws['A3'] = "Ticker:"
        ws['B3'] = results.get("ticker", "")
        ws['A4'] = "Company:"
        ws['B4'] = results.get("company_name", "")
        
        # Score
        score = results.get("score", {})
        ws['A6'] = "Overall Score:"
        ws['B6'] = score.get("total_score", 0)
        ws['B6'].font = Font(size=14, bold=True)
        
        # Recommendation
        ws['A7'] = "Recommendation:"
        ws['B7'] = results.get("recommendation", "")
        ws['B7'].font = Font(bold=True)
        
        # Strengths
        ws['A9'] = "Strengths:"
        ws['A9'].font = Font(bold=True)
        strengths = score.get("strengths", [])
        for i, strength in enumerate(strengths, start=10):
            ws[f'A{i}'] = f"• {strength}"
        
        # Concerns
        concern_row = 10 + len(strengths) + 1
        ws[f'A{concern_row}'] = "Concerns:"
        ws[f'A{concern_row}'].font = Font(bold=True)
        concerns = score.get("concerns", [])
        for i, concern in enumerate(concerns, start=concern_row + 1):
            ws[f'A{i}'] = f"• {concern}"
    
    def _write_metrics_sheet(self, ws, results: Dict[str, Any]):
        """Write detailed metrics to worksheet."""
        from openpyxl.styles import Font
        
        metrics = results.get("metrics", {})
        
        row = 1
        ws[f'A{row}'] = "Financial Metrics"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        row += 2
        
        # Write each category
        for category, values in metrics.items():
            ws[f'A{row}'] = category.replace("_", " ").title()
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            if isinstance(values, dict):
                for metric, value in values.items():
                    ws[f'A{row}'] = f"  {metric.replace('_', ' ').title()}"
                    ws[f'B{row}'] = f"{value:.2f}" if isinstance(value, (int, float)) else str(value)
                    row += 1
            
            row += 1
    
    def _write_scores_sheet(self, ws, results: Dict[str, Any]):
        """Write category scores to worksheet."""
        from openpyxl.styles import Font
        
        score = results.get("score", {})
        category_scores = score.get("category_scores", {})
        
        ws['A1'] = "Category Scores"
        ws['A1'].font = Font(size=14, bold=True)
        
        ws['A3'] = "Category"
        ws['B3'] = "Score"
        ws['A3'].font = Font(bold=True)
        ws['B3'].font = Font(bold=True)
        
        row = 4
        for category, value in category_scores.items():
            ws[f'A{row}'] = category.replace("_", " ").title()
            ws[f'B{row}'] = value
            row += 1
    
    def _write_simple_format(self, filepath: str, results: Dict[str, Any]) -> str:
        """Write simple text format as fallback."""
        with open(filepath, 'w') as f:
            f.write("=" * 80 + "\n")
            f.write("INVESTMENT ANALYSIS REPORT\n")
            f.write("=" * 80 + "\n\n")
            
            f.write(f"Ticker: {results.get('ticker', '')}\n")
            f.write(f"Company: {results.get('company_name', '')}\n")
            f.write(f"Score: {results.get('score', {}).get('total_score', 0)}\n")
            f.write(f"Recommendation: {results.get('recommendation', '')}\n\n")
            
            score = results.get("score", {})
            f.write("STRENGTHS:\n")
            for strength in score.get("strengths", []):
                f.write(f"  • {strength}\n")
            
            f.write("\nCONCERNS:\n")
            for concern in score.get("concerns", []):
                f.write(f"  • {concern}\n")
            
            f.write("\n" + "=" * 80 + "\n")
        
        return filepath
